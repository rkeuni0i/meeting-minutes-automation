"""
회의록 자동화 도구
엑셀로 정리된 회의 내용을 읽어 OpenAI API로 요약하고,
스타일이 적용된 엑셀 파일로 다시 저장한다.
"""

import os
import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 0. 환경변수 로드 (.env 파일에서 OPENAI_API_KEY 읽어옴) ─────────
load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

INPUT_PATH = "data/회의록.xlsx"
OUTPUT_PATH = "output/AI_회의록_정리본.xlsx"

SYSTEM_PROMPT = """
너는 회의록 작성 전문가이다.
아래 회의 내용을 분석해서 반드시 아래 형식으로 작성해줘.
회의 목적:
주요 논의 사항:
결정 사항:
담당자별 업무:
추가 확인 사항:
다음 회의 안건:
각 항목은 반드시 줄바꿈해서 작성하고,
항목명 뒤에 ':'을 붙여 작성해줘.
'다음 회의 안건' 항목은 여러 개일 경우 각 안건을 '- '로 시작하는
별도의 줄로 작성해줘. (예: - 예산안 재검토)
"""


def style_sheet(ws, df):
    """엑셀 시트에 헤더 색상, 테두리, 열 너비 등 스타일을 적용한다."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col_idx, col_name in enumerate(df.columns, start=1):
        if len(df) > 0:
            max_len = max(
                [len(str(col_name))] +
                [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)]
            )
        else:
            max_len = len(str(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 5, 60)

    ws.freeze_panes = "A2"


def load_meeting_text(path: str) -> str:
    """엑셀에서 회의 내용 컬럼을 읽어 하나의 텍스트로 합친다."""
    df = pd.read_excel(path)
    return "\n".join(df["회의내용"])


def summarize_meeting(meeting_text: str) -> str:
    """OpenAI API를 호출해 회의 내용을 정해진 형식으로 요약한다."""
    response = client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": meeting_text},
        ],
    )
    return response.choices[0].message.content


def parse_summary(summary: str):
    """GPT가 반환한 요약 텍스트를 회의록 / 담당자별 To-do / 다음 회의 안건으로 분리한다."""
    lines = summary.split("\n")
    meeting_data = []
    agenda_list = []
    category = ""

    for line in lines:
        line = line.strip()
        if not line:
            continue

        if line.startswith("-") and category == "다음 회의 안건":
            agenda_list.append(line.lstrip("-").strip())
            continue

        if ":" in line:
            category, content = line.split(":", 1)
            category = category.strip()
            content = content.strip()

            if category == "다음 회의 안건":
                if content:
                    agenda_list.append(content)
                continue

            meeting_data.append([category, content])
        else:
            if meeting_data:
                if line.startswith("-"):
                    meeting_data[-1][1] += "\n" + line
                else:
                    meeting_data[-1][1] += " " + line

    summary_df = pd.DataFrame(meeting_data, columns=["구분", "내용"])

    if not agenda_list:
        agenda_list = ["다음 회의 안건 없음"]

    agenda_df = pd.DataFrame({
        "번호": range(1, len(agenda_list) + 1),
        "안건": agenda_list,
    })

    return summary_df, agenda_df


def extract_todos(summary_df: pd.DataFrame) -> pd.DataFrame:
    """'담당자별 업무' 행에서 담당자별 To-do 목록을 추출한다."""
    todo_row = summary_df[summary_df["구분"] == "담당자별 업무"]
    todo_list = []

    if not todo_row.empty:
        todo_text = todo_row["내용"].values[0]
        for part in todo_text.split(","):
            part = part.strip()
            if ":" in part:
                name, task = part.split(":", 1)
                todo_list.append([name.strip(), task.strip()])

    if not todo_list:
        todo_list = [["-", "담당자별 업무 없음"]]

    return pd.DataFrame(todo_list, columns=["담당자", "업무"])


def save_to_excel(summary_df, todo_df, agenda_df, output_path: str):
    """정리된 결과를 시트 3개(회의록/To-do/다음 안건)로 나누어 엑셀로 저장한다."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="회의록", index=False)
        todo_df.to_excel(writer, sheet_name="담당자별 To-do", index=False)
        agenda_df.to_excel(writer, sheet_name="다음 회의 안건", index=False)

        style_sheet(writer.sheets["회의록"], summary_df)
        style_sheet(writer.sheets["담당자별 To-do"], todo_df)
        style_sheet(writer.sheets["다음 회의 안건"], agenda_df)


def main():
    meeting_text = load_meeting_text(INPUT_PATH)
    summary = summarize_meeting(meeting_text)
    summary_df, agenda_df = parse_summary(summary)
    todo_df = extract_todos(summary_df)
    save_to_excel(summary_df, todo_df, agenda_df, OUTPUT_PATH)
    print(f"저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
